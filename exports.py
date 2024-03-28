import openpyxl
from openpyxl import load_workbook
from datetime import datetime, timedelta
from datetime import time
from dateutil import parser
import json
import re
import custom_exceptions as ex

import payroll as pr

EXPORT_DURATION = 28 #days
def exportStudy(studyDict, stateVersion, saveFilePath=None):
	debug = False
	wb = openpyxl.Workbook()
	sheet = wb.active

	print("EXPORTING ", studyDict)

	descList = []
	if stateVersion == "NT":
		descList = list(pr.DESCRIPTORS_SHIFTS_ALLOTHERS_NT.values())
		descList.extend(list(pr.DESCRIPTORS_SHIFTS_PENS_NT.values()))
		descList.extend(pr.DESCRIPTORS_OTHER_NT)
	elif stateVersion == "WA":
		descList = list(pr.DESCRIPTORS_SHIFTS_ALLOTHERS_WA.values())
		descList.extend(list(pr.DESCRIPTORS_SHIFTS_PENS_WA.values()))
		descList.extend(pr.DESCRIPTORS_OTHER_WA)
	descList = list(set(descList))
	descList.sort()


	descDict = {} #Stores the row that this description is located in.
	start = 2
	for desc in descList:
		sheet["A"+str(start)] = desc
		descDict.update({desc:start}) #filling descDict
		start += 1
	if debug:
		print(f'Desc dict: {descDict}')


	start = 3
	dateDict = {}
	tempList = list(studyDict.keys())
	startDate = datetime.strptime(tempList[0], "%d-%m-%Y")
	endDate = datetime.strptime(tempList[-1], "%d-%m-%Y")
	while startDate <= endDate:
		sheet.cell(row=1,column=start).value = startDate.strftime("%d-%m-%Y")
		dateDict.update({startDate.strftime("%d-%m-%Y"):start})
		startDate += timedelta(days=1)
		start += 1
	if debug:
		print(f'dates dict: {dateDict}')

	for date, hoursList in studyDict.items():
		for hoursType in hoursList:
			try:
				daRow = descDict[hoursType["description"].strip()]
			except KeyError as e:
				print(f"Key error: {e} - SKIPPING THIS IN THE EXPORT!")
				continue
				# return False, f"Descriptions don't match current state setting of '{stateVersion}'"
			try:
				daCol = dateDict[date]
			except KeyError as e:
				print(f"Key error: {e}")
				return False, f"Dates out of range - '{e}'"
			try:
				daVal = float(hoursType["units"])
			except KeyError as e:
				daVal = 0
				# Don't care if it cant find a quantity, just put zero. Probably doesn't matter anyway.

			sheet.cell(row=daRow, column=daCol).value = daVal

	if not saveFilePath:
		print("USING DEFAULT FILE NAME")
		saveFilePath = "./export-" + datetime.now().strftime("%Y-%m-%d-%H%M") + ".xlsx"
	# if not re.search(r'^\./', saveFilePath):
	# 	raise ValueError("File Path not in correct format")
	wb.save(saveFilePath) #No real way of knowing if this was successful
	return True, "Saved successfully"

