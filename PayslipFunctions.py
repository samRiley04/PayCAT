import PyPDF2
import openpyxl
from openpyxl import load_workbook
import yaml
from datetime import datetime, timedelta
from datetime import time
from dateutil import parser
import json
import re
import custom_exceptions as ex

#Alerts - in the form [priority (HI/MED/LO), title, body]
alerts = []

#Descriptions dict - to aid with payslip ingestion. List of known possible descriptions.
descShortlist = [
	"BASE HOURS",
	"PUBLIC HOLIDAY 1.5",
	"PROFESSIONAL DEVT ALLOW",
	"SMART SALARY SP FIXED",
	"PENALTIES AT 25%",
	"PENALTIES AT 20%",
	"PENALTIES AT 75%",
	"OVERTIME @ 1.5",
	"MEAL - BREAKFAST MED PRAC",
	"MEAL - DINNER MED PRACT",
	"ANNUAL LEAVE"
]

# Dynamically generated if the roster utilises shift-codes (where instead a straight "0800-1800" listed in the roster, they use a code "AM" which is then referenced in a table elsewhere to mean those hours.)
SHIFT_CODES = {}
ATTEMPTS = []


def isValidShiftCode(string):
	# FOR NOW - only requirement is to start with a letter.
	if isinstance(string, str):
		if string.upper() in ["OFF", "RDO"]:
			return False #Should ignore these strings as they definitely aren't shift codes, but technically fit the description and are used sometimes.
		try:
			return (re.search(r'^\w+$', string)) #Match only a single run of letters/numbers, no spaces, dashes etc.
		except (AttributeError, IndexError):
			pass
	return False

# Returns the number of hours represented by a string of a given time range
# e.g. '0800-1800' returns 10
def parseHours(string):
	s = None
	try:
		# This regular expression makes me aroused
		# accepts (HHHH or (1 to 9)HH or 000)-(HHHH or (1 to 9)HH or 000)
		# or
		# accepts (HH:HH or (1 to 9):HH or 0:00)-(HH:HH or (1 to 9):HH or 0:00)
		s = re.search(r'(\d{2}(:?)\d{2}|[1-9](:?)\d{2}|0(:?)00)\s*-\s*(\d{2}(:?)\d{2}|[1-9](:?)\d{2}|0(:?)00)', string)
	except TypeError:
		pass
	if not s is None:
		times = s.group().replace(' ','').replace(":","").split("-")
		outputStr = ""
		# Have to clone the array else it gives errors.
		for indx, t in enumerate(times.copy()):
			# Regex has ensured this is safe to do - if time has no leading zero, add it. (required for datetime parsing)
			if len(t) == 3:
				times[indx] = "0"+times[indx]
		#May not be the most elegant way to have done this but so be it.
		return times[0] + "-" + times[1]
	else:
		return None

ROSTERS_YEAR = None

# Returns True if roster_year is set. If not set, finds it.
def findYear(sheet):
	if ROSTERS_YEAR is None:
		for row in sheet.iter_rows():
			for cell in row:
				if not cell.value is None:
					tryDate = parseDate(str(cell.value))
					if tryDate:
						ROSTERS_YEAR = int(datetime.strptime(tryDate, "%d-%m-%Y").strftime("%Y"))
	print(ROSTERS_YEAR)
	print("!@#!@#!123!@#!@#")
	return True


def parseDate(string):
	s = None
	try:
		# YYYY-MM-DD or XX-XX-XX or DD-MM-YYYY
		s = re.search(r'(\d{4}[-/\. ]\d{2}[-/\. ]\d{2})|(\d{2}[-/\. ]\d{2}[-/\. ](\d{4}|\d{2}))', string)
	except TypeError:
		pass
	if not s is None:
		return s.group()
	else:
		return None

def findShiftCodes(sheet, debug):
	if not ATTEMPTS == []:
		return False
	ATTEMPTS.append("one")	
	for row in sheet.iter_rows():
		for cell in row:
			nextCell = sheet.cell(row=cell.row, column=cell.column+1)
			if isValidShiftCode(cell.value) and parseHours(nextCell.value):
				SHIFT_CODES.update({str(cell.value):parseHours(nextCell.value)})
	if debug:
		print("SHIFT CODES:")
		print(json.dumps(SHIFT_CODES, indent=2))
	return True


def ingestTypeA(sheet, findName, debug):
	outputDict = {}
	tempDates = {}
	# find DATE CELLS
	# AND check validity of DATES simultaneously
	for row in sheet.iter_rows():
		for cell in row:
			# If the cell isn't empty
			if not (cell.value is None):
				dateAttempt = parseDate(str(cell.value))
				#parseDate returns None if it doesn't match a date.
				# Then check if it's a valid bonafide date.
				if not (dateAttempt is None):
					if dateValidTypeA(cell, sheet):
						# Copy whole cell object so can access row/col easier.
						tempDates.update({
							str(dateAttempt):cell
						})
					elif debug:
						print("DISCARDING potential date: " + dateAttempt)
	if len(tempDates) == 0:
		#Found no recognisable dates in the roster
		raise ex.NoRecognisedDates()
	if debug:
		print("tempDates: "+ str(tempDates))
	# FIND HOURS COL
	tempHoursCounts = {} 
	# Gather some candidates for potential shift-columns
	for col in sheet.iter_cols():
		for cell in col:
			if parseHours(cell.value):
				if cell.column_letter not in tempHoursCounts:
					tempHoursCounts.update({cell.column_letter:1})
				else:
					tempHoursCounts.update({cell.column_letter:(tempHoursCounts[cell.column_letter]+1)})
	if tempHoursCounts == {}:
		# Did not find a name.
		raise ex.NoRecognisedShifts()
	# NOW pick only the column with most hours
	hoursCol = list(tempHoursCounts.keys())[0] #Pick any, the first element will do.
	for candidateCol, count in tempHoursCounts.items():
		if count > tempHoursCounts[hoursCol]:
			hoursCol = candidateCol
	#hoursCol is now your winner
	if debug:
		for k, v in tempHoursCounts.items():
			print("{k} - {v}".format(k=k, v=v))
		print("column \'"+hoursCol+"\' wins!")
	for dateKey, dateCell in tempDates.items():
		# Find names.
		for col in sheet.iter_cols(min_col=dateCell.column, max_col=dateCell.column, min_row=dateCell.row+1):
			for cell in col:
				if parseDate(str(cell.value)) in tempDates:
					break
					# We have reached the next layer of dates, so should stop looking for now (otherwise we will bleed into adjacent weeks.)
				if (not cell.value is None) and (cell.value == findName or re.search(r'[(\[{]'+re.escape(findName)+r'[)\]}]\s*$', str(cell.value))):
					# IF the cell contains your name but just after some other text, include that also.
					hrs = parseHours(sheet[hoursCol+str(cell.row)].value)
					if (hrs is None) and isValidShiftCode(cellVal): #parseHours couldn't find valid hours, check if it could be a shift code
						if SHIFT_CODES == {}:
							print("finding SHIFT CODES")
							findShiftCodes(sheet, debug)
						try:
							hrs = SHIFT_CODES[cellVal]
						except (KeyError):
							pass
							print("couldn't find that key: "+cellVal)
					if not (hrs is None):
						# Only update the dict if hours worked is a value
						outputDict.update({dateKey:hrs})
	if debug:
		print("Pre-cull outputDict: ")
		print(json.dumps(outputDict, indent=4))
	return outputDict


def ingestTypeB(sheet, findName, debug=True):
	outputDict = {}
	tempDates = {}
	# find DATE CELLS
	# AND check validity of DATES simultaneously
	for row in sheet.iter_rows():
		for cell in row:
			# If the cell isn't empty
			if not (cell.value is None):
				dateAttempt = parseDate(str(cell.value))
				#parseDate returns None if it doesn't match a date.
				# Then check if it's a valid bonafide date. (check down first, if that fails, laterally)
				if not (dateAttempt is None):
					if dateValidTypeB(cell, sheet):
						# Copy whole cell object so can access row/col easier.
						tempDates.update({
							str(dateAttempt):cell
						})
					elif debug:
						print("DISCARDING potential date: " + dateAttempt)
	if len(tempDates) == 0:
		#Found no recognisable dates in the roster
		raise ex.NoRecognisedDates()
	if debug:
		print("tempDates: "+ str(tempDates))
	# find NAME COLS
	tempNameRows = []
	for col in sheet.iter_cols():
		for cell in col:
			if cell.value == findName:
				tempNameRows.append(cell.row)
	if len(tempNameRows) == 0:
		# Did not find a name.
		raise ex.NameNotFound()
	# create outputDict
	# Remember tempDates contains {DATES:cells}
	if debug:
		print("tempNameRows: " + str(tempNameRows))
	for dateKey in tempDates:
		# Find out if the nameRow is the closest row AFTER the date in question (prevent doubling up due to vertical roster stacking)
		closestRow = max(tempNameRows)
		for row in tempNameRows:
			rowDif1 = row - int(tempDates[dateKey].row)
			rowDif2 = closestRow - int(tempDates[dateKey].row)
			# >0 ensures the row is AFTER the date in question.
			if (rowDif1 < rowDif2) and (rowDif1 > 0):
				closestRow = row
			#if debug:
				#print("attempted row: " + str(row)+". dif was "+str(rowDif1) + " and now closest row is " + str(closestRow))
		if debug:
			print("col: "+ tempDates[dateKey].column_letter + " | row: "+ str(closestRow))
		cellVal = sheet[str(tempDates[dateKey].column_letter)+str(closestRow)].value
		hrs = parseHours(cellVal)
		if (hrs is None) and isValidShiftCode(cellVal): #parseHours couldn't find valid hours, check if it could be a shift code
			if SHIFT_CODES == {}:
				print("finding SHIFT CODES")
				findShiftCodes(sheet, debug)
			try:
				hrs = SHIFT_CODES[cellVal]
			except (KeyError):
				pass
				print("couldn't find that key: "+cellVal)
		if not (hrs is None):
			# Only update the dict if hours worked is a value
			outputDict.update({dateKey:hrs})
	if debug:
		print("Pre-cull outputDict: ")
		print(json.dumps(outputDict, indent=4))
	return outputDict

def ingestTypeC(sheet, findName, debug):
	outputDict = {}
	tempNameCols = []
	# find NAME CELLS
	# AND check validity of NAMES simultaneously
	for row in sheet.iter_rows():
		for cell in row:
			# If the cell isn't empty
			if not (cell.value is None) and (cell.value == findName) and not (cell.column_letter in tempNameCols):
				tempNameCols.append(cell.column_letter)
	if len(tempNameCols) == 0:
		#Found no recognisable name in the roster
		raise ex.NameNotFound()
	if debug:
		print("tempNameCols: " + str(tempNameCols))
	# find DATE ROWS (include the dates as well {"2":"2023-08-22"})
	tempDates = {}
	for col in sheet.iter_cols():
		for cell in col:
			if not (cell.value is None):
				dateAttempt = parseDate(str(cell.value))
				if not (dateAttempt is None):
					if dateValidTypeC(cell, sheet):
						tempDates.update({
							cell.row:dateAttempt
							})
					elif debug:
						print("DISCARDING potential date: " + dateAttempt)
	if len(tempDates) == 0:
		raise ex.NoRecognisedDates()
	if debug:
		print("tempDates: " + str(tempDates))
	# create outputDict
	# Remember tempNameCols contains [col, col]
	# Iterate through each name, then collect all the date-rows below it
	for col in tempNameCols:
		for row in tempDates:
			cellVal = sheet[str(col)+str(row)].value
			if cellVal is None:
				continue
			hrs = parseHours(cellVal)
			if (hrs is None) and isValidShiftCode(cellVal): #parseHours couldn't find valid hours, check if it could be a shift code
				findShiftCodes(sheet, debug) #will auto-cancel if it's already been done.
				try:
					hrs = SHIFT_CODES[cellVal]
				except (KeyError):
					pass
					if debug:
						print("couldn't find that key: "+cellVal)
			if not (hrs is None):
				# Only update the dict if hours worked is a value
				outputDict.update({
					tempDates[row]:hrs
				})
	return outputDict

# SENSITIVITY MAY NEED TO BE TITRATED TO EFFECT in future.
def dateValidTypeA(cell, sheet):
	debug = False
	if cell.column == 1:
		compareLeft = "<INVALID_DATE>"
	else:
		compareLeft = str(sheet.cell(row=cell.row, column=cell.column-1).value)
	compareRight = str(sheet.cell(row=cell.row, column=cell.column+1).value)
	if debug:
		print("R: ")
		print(compareRight)
		print("L: ")
		print(compareLeft)
	# If either of left or right are valid dates.
	if (parseDate(compareLeft) or parseDate(compareRight)):
		return True
	return False

def dateValidTypeB(cell, sheet):
	debug = False
	counter = 0
	# After marginValue number of cells that dont contain valid strings, consider this date INVALID.
	marginValue = 4
	# CHECK FOR HOURS BELOW
	# Create a range and iterate (next #marginValue cells not including the date cell itself.)
	for rowDown in sheet[cell.column_letter+str(cell.row+1):cell.column_letter+str(cell.row+marginValue)]:
		for cellDown in rowDown:
			# counter += 1
			# if debug:
			# 	print("date: " + str(cell.value) + " | checking: " + str(cellDown.value))
			# if not (parseHours(cellDown.value) is None):
			# 	# Got at least one valid string before the marginValue cutoff, so it's Valid.
			# 	return True
			# elif not (cellDown.value is None):
			# 	# If you encounter any non-hours-like shit below, it's likely invalid
			# 	return False
			# elif counter >= marginValue:
				# Unclear if invalid (only saw empty cells)
				# CHECK FOR DATES EITHER SIDE
				# Create a range and iterate
			if cell.column == 1:
				compareLeft = "<INVALID_DATE>"
			else:
				compareLeft = str(sheet.cell(row=cell.row, column=cell.column-1).value)
			compareRight = str(sheet.cell(row=cell.row, column=cell.column+1).value)
			if debug:
				print("R: ")
				print(compareRight)
				print("L: ")
				print(compareLeft)
			# If either of left or right are valid dates.
			if (parseDate(compareLeft) or parseDate(compareRight)):
				return True
			return False

def dateValidTypeC(cell, sheet):
	debug = False
	"""
	<TEMPORARILY DISABLED to titrate sensitivity - to be considered for re-inclusion in the future if indicated>
	counter = 0
	marginValue = 4
	# CHECK FOR HOURS ON RIGHT
	# Create a range and iterate (next #marginValue cells not including the date cell itself.)
	rangeStart = sheet.cell(column=(cell.column + 1), row=cell.row).coordinate
	rangeEnd = sheet.cell(column=(cell.column + marginValue), row=cell.row).coordinate
	for colsAcross in sheet[rangeStart:rangeEnd]:
		for cellAcross in colsAcross:
			counter += 1
			if debug:
				print("date: " + str(cell.value) + " | checking: " + str(cellAcross.value))
			if not (parseHours(cellAcross.value) is None):
				# Got at least one valid string before the marginValue cutoff, so it's Valid.
				return True
			elif not (cellAcross.value is None):
				# If you encounter any non-hours-like shit below, it's likely invalid
				return False
			elif counter >= marginValue:
				# Unclear if invalid (only saw empty cells)
			"""
				# CHECK FOR DATES ABOVE/BELOW
				# Create a range and iterate
	if cell.row == 1:
		compareAbove = "<INVALID_DATE>"
	else:
		compareAbove = str(sheet.cell(row=cell.row-1, column=cell.column).value)
	compareBelow = str(sheet.cell(row=cell.row+1, column=cell.column).value)
	if debug:
		print("AB: ")
		print(compareAbove)
		print("BEL: ")
		print(compareBelow)
	# If either of above or below are valid dates.
	if (parseDate(compareAbove) or parseDate(compareBelow)):
		return True
	return False

# RETURN ERRORS
# ValueError - "employee name not found", "recognisable dates not found"
def ingestRoster(fileName, findName, rosterFormat, startDate, endDate, ignoreHidden=True, debug=False):
	debug=True
	outputDict = {}
	wb = load_workbook(fileName, data_only=True)
	sheet = wb.active
	try:
		if rosterFormat == "A":
			outputDict = ingestTypeA(sheet, findName, debug)
		elif rosterFormat == "B":
			outputDict = ingestTypeB(sheet, findName, debug)
		elif rosterFormat == "C":
			outputDict = ingestTypeC(sheet, findName, debug)
		else:
			#the fuck?
			raise ValueError()
	except (ValueError):
		raise ValueError("Invalid roster format: " + str(rosterFormat))
	except (ex.NameNotFound):
		raise ValueError("Found no recognisable name in the roster \'"+fileName+"\'. Is it spelled correctly?")
	except (ex.NoRecognisedDates):
		# This is usually suggestive of the wrong roster format being provided!
		raise ValueError("Found no recognisable dates in the roster '{fileName}'. Did you select the correct roster format?".format(fileName=fileName.split("/")[-1]))
	except (ex.NoRecognisedShifts):
		raise ValueError(f"Could not find any valid shift-hours in this roster {fileName.split('/')[-1]}")
	#Now all the roster is ingested, trim the dict using the start/end dates.
	if debug:
		print("output dict:")
		print(outputDict)
		print("start: " + str(startDate) + ", end: " + str(endDate))
	copy = outputDict.copy()
	for entry in copy:
		e = datetime.strptime(entry, "%Y-%m-%d")
		if not (e >= startDate and e <= endDate):
			outputDict.pop(entry)
	if debug:
		print("output dict:")
		print(json.dumps(outputDict,indent=4))
	if outputDict == {}:
		raise ValueError("Found no recognisable dates in the roster '{fileName}' for the given date range {sD} to {eD}.".format(fileName=fileName.split("/")[-1], sD=startDate.date(), eD=endDate.date()))
	return outputDict


#print(json.dumps(ingestRoster("TESTING/OPH.xlsx", "Samuel Riley", "C", datetime.strptime("2023-02-13", "%Y-%m-%d"), datetime.strptime("2023-02-26", "%Y-%m-%d")), indent=4))


# TODO - re-write this entire function using Regex (will be a LOT shorter). I am dumb.
def ingestPDF(fileName):
	payPeriodLength = 14
	#Payslip Dictionary (the end product of ingesting the payslip)
	psDict = {}
	with open(fileName, 'rb') as pdf:
		pdfReader = PyPDF2.PdfReader(pdf)
		try:
			pageOfInterest = pdfReader.pages[1].extract_text().split('\n')
		except IndexError:
			raise ValueError("This PDF doesn't seem to fit the format of a WA health payslip.")
		LINELIMITER = 0
		# PAGE TWO ----
		for line in pageOfInterest:
			#If this line starts with a number (the only lines we are interested in do)
			if line[0][0].isnumeric():
				#if LINELIMITER <= 9:
				#	LINELIMITER+=1
				#else:
				#	break
				#print('DOING-------')
				# Remove all those random double sapces
				wordsList = " ".join(line.split()).split(" ")
				#print(wordsList)
				date = ""
				description = ""
				units = ""
				rate = ""
				amount = ""

				# GO WORD BY WORD until we get our description done (requires iterating)
				for index, word in enumerate(wordsList):
					if not description == "":
						# Once we have our description, we don't need to iterate on the list anymore.
						break

					# ?Date (and we don't already have a date stored)
					if word.replace('-','').isnumeric() and date == "":
						date = str(word)
					# ?Description
					# This one is harder
					elif word.isalpha():
						# If we already have a description, pass on.
						if not description == "":
							continue
						# Attempt to consolidate the word using known list of descriptions.
						# Iterate through the remainder of the words from where we are up to, creating a franken-word to see if it fits any known descriptions.
						frankenword = ""
						for candidate in wordsList[index:]:
							# Attempt to identify if we have reached the "units" column erroneously (check the last three charcters)
							frankenword += candidate
							if frankenword in descShortlist:
								description = frankenword
								break
							else:
								frankenword += " "
							#If we have reached and checked a number and the word still doesn't match, we have gone too far. Subtract the number and just go with the words part, must be a new term.
							#Checking for numbers AFTER the frankenword ensures that descriptions CONTAINING A NUMBER will still be stored correctly.
							#It makes having to remove the string we just added worth it.
							if candidate.replace(".","").isnumeric():
								description = frankenword[:-len(candidate)-1]
								alerts.append(["MED", "New Description", "The description \'"+description+"\' has not been encountered before. Should we record it into the dictionary?"])
								break
				# NOW COLLECT the numeric values.
				# (but first do this) If the date-entry doesn't already exist (it might), initialise it as a list.
				if not date in psDict.keys():
					psDict[date] = []

				# Two formats of payslip rows:
				# (1) DATE - DESCRIPTION - UNITS - RATE - AMOUNT
				# (2) DATE - SECOND DATE - DESCRIPTION - AMOUNT
				# Use this information to trickily collect the last data and then input it into the dict.
				if wordsList[1].count("-") == 2:
					amount = wordsList[-1].replace(",", "") #remove commas
					psDict[date].append({"description":description,
									"amount":amount})
				else:
					units = wordsList[-3]
					rate = wordsList[-2]
					amount = wordsList[-1]
					psDict[date].append({"description":description,
									"units":units,
									"rate":rate,
									"amount":amount})
		
		# NOW grab all the random info from page 1 and fill the psDict.
		
		# PAGE ONE ----
		pageOfInterest = pdfReader.pages[0].extract_text().split('\n')
		employer = pageOfInterest[0] #Seems to be the rule in this case.
		employeeName = ""
		payPeriodEnding = ""
		totalPretaxIncome = ""
		for index, line in enumerate(pageOfInterest):
			if employeeName == "" and "Name: " in line:
				words = line.split(' ')
				for word in words[1:]: #Cut out the first word as it will be "Name: "
					if word == "Employee" :
						#once we reach the end of the name, break.
						break
					employeeName += word + " "
				employeeName.strip() 
			elif totalPretaxIncome == "" and "3. TOTAL TAXABLE EARNINGS" in line:
				words = pageOfInterest[index+1].strip().split(" ")
				totalPretaxIncome = words[0].replace(",","")
			elif payPeriodEnding == "" and "Period End Date: " in line:
				words = line.split(' ')
				payPeriodEnding = words[-1]

		# Calculate the start of the pay period.
		ppStart = datetime.strptime(payPeriodEnding, "%d-%m-%Y") - timedelta(days=payPeriodLength)

		fullDict = {
			"name":fileName.split("/")[-1],
			"employeeName":employeeName,
			"employer":employer,
			"totalPretaxIncome":totalPretaxIncome,
			"payPeriodStart":ppStart.strftime("%d-%m-%Y"),
			"payPeriodEnding":payPeriodEnding,
			"data":psDict
		}
		#print(json.dumps(fullDict, indent=4))
		return fullDict

	#print("-------")

def neatenDate(string, earliestDate=None):
	if earliestDate and ("_NO_YEAR_YET_" in string):
		return string.replace("_NO_YEAR_YET_", earliestDate.strftime("%Y"))
		
	year = "_NO_YEAR_YET_"
	if earliestDate:
		year = earliestDate.strftime("%Y")

	findHalfDate = re.search(r'^\d{1,2}[-\/]\d{1,2}$', string)
	if findHalfDate:
		if earliestDate:
			return findHalfDate.group().replace("/", "-") + "-" + earliestDate.strftime("%Y")
		else:
			return findHalfDate.group().replace("/", "-") + "-_NO_YEAR_YET_"
	findFullDate = re.search(r'^\d{1,2}[-\/]\d{1,2}[-\/]\d{4}$', string)
	if findFullDate:
		return findFullDate.group().replace("/", "-")
	
	return None

def neatenUnits(string):
	toReturn = 0
	hrs = re.search(r'\d+H', string)
	try:
		toReturn += int(hrs.group()[:-1]) # cut off the 'H'
	except (IndexError, AttributeError):
		pass
	mins = re.search(r'\d+M', string)
	try:
		toReturn += float(mins.group()[:-1])/60
	except (IndexError, AttributeError):
		pass

	if toReturn == 0:
		return None

	return str(toReturn)

NT_EMPLOYEE_NAME_RE = r'^Name: '
NT_EMPLOYER_RE = r'Employer(\s*)ABN(\s*)Name:'
NT_PTI_RE = r'Total(\s*)Gross:'
NT_PPE_RE = r'Pay(\s*)Period(\s*)From/To:'

NT_DESC_SHORTLIST = [
	"Perishable Allowance",
	"PROF DEVELOPMENT AST",
	"OT",
	"Shiftduty",
	"Salary Sacrifice",
	"Salary Payment"
]

def ingestPayslip(fileName, version="WA", debug=True):
	earliestDate = None
	payPeriodLength = 14

	psDict = {}
	employeeName = None
	employer = None
	preTaxIncome = None
	payPeriodEnding = None

	defaultDate = datetime.strptime("01-01-9001", "%d-%m-%Y")

	with open(fileName, 'rb') as pdf:
		pdfReader = PyPDF2.PdfReader(pdf)
		if version == "WA":
			return ingestPDF(fileName)
			# TODO make this apply to WA.
		elif version == "NT":
			lines = []
			for x in pdfReader.pages:
				lines.extend(x.extract_text().split('\n'))
			# VALIDATION
			for x in lines:
				look = re.search("NORTHERN TERRITORY GOVERNMENT", x)
				if look:
					break
			else:
				raise ValueError("This doesn't look like an NTG payslip." )

			for index, line in enumerate(lines):
				print(line)

				temp = re.search(NT_PPE_RE, line)
				if not payPeriodEnding and temp:
					print("found ", temp.group())
					dates = re.findall(r'\d{2}\s*\w{3}\s*\d{4}', line[temp.span()[1]:])
					payPeriodEnding = dates[-1]
					if not earliestDate:
						earliestDate = datetime.strptime(dates[0].strip(), "%d %b %Y")
					continue

				temp = re.search(NT_EMPLOYEE_NAME_RE, line)
				if not employeeName and temp:
					employeeName = line[temp.span()[1]:].strip()
					print("EMPLOYEE NAME ", employeeName)
					continue

				temp = re.search(NT_EMPLOYER_RE, line)
				if not employer and temp:
					employer = line[temp.span()[1]:].strip()
					continue

				temp = re.search(NT_PTI_RE, line)
				if not preTaxIncome and temp:
					preTaxIncome = line[temp.span()[1]:].strip().replace("$", "").replace(",", "")
					continue

				# If the line starts with a known shift description
				# Assumption - descriptions don't contain numbers.
				shiftDesc = re.search(r'^[A-Za-z]+\s([A-Za-z]+\s)*', line) # Find first sequential words
				print("---", shiftDesc)
				if shiftDesc and (shiftDesc.group().strip() in NT_DESC_SHORTLIST):
					# Search for sequential lines that may have be inappropriately split
					for x in range(index+1, index+3):
						if x+1 >= len(lines):
							break # reached the end
						line = line.replace(",", "")

						check = re.search(r'^[A-Za-z]+\s([A-Za-z]+\s)*', lines[x])	#SHOULDN'T start with a word (should start with @ or digits)
						if check:
							break #stop and do no more, else risk overflowing into new shift entries.

						check = re.search(r'@?\s*\d+\.\d{2}\s*(P\/H)?\s*-?\$\d+\.\d{2}', lines[x]) #the rate/amount match string
						if check:
							line += check.group() #usually this will only be called once or twice
							continue
						check = re.search(r'\d+M', lines[x]) #the minutes match string
						if check:
							line += check.group() #usually this will only be called once or twice
							continue
					print("---", line)

					description = shiftDesc.group().strip()

					date = None
					temp2 = re.search(r'\d{1,2}[-\/]\d{1,2}([-\/]\d{4})?', line) # DD/MM or DD/MM/YYYY match string
					if temp2:
						date = neatenDate(temp2.group().strip())

					units = None
					temp2 = re.search(r'(\d+H)(\s*(\d+M))?', line)
					if temp2:
						print("--- UNITS: ", temp2.group())
						units = neatenUnits(temp2.group().strip())
						print("--- UNITS: ", units)

					rate = None
					temp2 = re.search(r'@\s*\d+\.\d{2}', line)
					if temp2:
						rate = temp2.group().strip().replace("@", "")

					amount = None
					temp2 = re.findall(r'-?\$\d+\.\d{2}', line)
					if temp2:
						print("TEMP2: ", temp2)
						temp2 = list(temp2)[-1] #pick the last match
						amount = temp2.strip().replace("$", "")

					if not date:
						date = defaultDate.strftime("%d-%m-%Y")
						defaultDate += timedelta(days=1)
						if debug:
							print("PAYSLIP --- creating a fake date as we could not find one. For: ", line)
					date.replace("/", "-")
					print(date)
					if date not in psDict:
						psDict.update({
							date:[]
							})
					makeDict = {
						"description":description
					}
					if units: 
						makeDict.update({"units":units})
					if rate: 
						makeDict.update({"rate":rate})
					makeDict.update({"amount":amount})	
					psDict[date].append(makeDict)

			if earliestDate:
				for key, value in psDict.copy().items():
					if "_NO_YEAR_YET_" in key:
						psDict.pop(key)
						psDict.update({
							neatenDate(key, earliestDate): value
						})
			# SORT psDict
			keysList = []
			for x in psDict:
				keysList.append(datetime.strptime(x, "%d-%m-%Y"))
			sortedStuff = {}
			for y in sorted(keysList):
				ki = y.strftime("%d-%m-%Y")
				sortedStuff.update({
					ki:psDict[ki]
				})
			psDict = sortedStuff
			print(json.dumps(psDict, indent=2))
			ppStart = datetime.strptime(payPeriodEnding, "%d %b %Y") - timedelta(days=payPeriodLength)
			fullDict = {
				"name":fileName.split("/")[-1],
				"employeeName":employeeName,
				"employer":employer,
				"totalPretaxIncome":preTaxIncome,
				"payPeriodStart":ppStart.strftime("%d-%m-%Y"),
				"payPeriodEnding":datetime.strptime(payPeriodEnding, "%d %b %Y").strftime("%d-%m-%Y"),
				"data":psDict
			}
			return fullDict
	

#print(json.dumps(ingestPDF("test.pdf"), indent=4))
